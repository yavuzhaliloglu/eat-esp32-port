#!/usr/bin/env python3
"""
Saha testi icin surekli olcum kaydedici.

Ne yapiyor:
  1) Bilgisayardaki seri portlari tarar, her birinde gercek bir sayac olup
     olmadigini dener (readout-mode.py'deki AYNI kimlik dogrulama protokolu
     ile), sonucu bir liste halinde gosterip hangisini kullanacagini sorar.
  2) Kac dakikada bir olcum alinacagini sorar.
  3) O periyotla surekli okuma yapip cihazin verdigi HER ALANI (duz okuma/
     kimlik yaniti, kisa okuma ozet alanlari, uzun okuma esik+reset kayitlari
     - hepsi) bir Excel (.xlsx) dosyasina TEK TEK, alt alta satir olarak
     ekler - "PC Tarih/Saat | Okuma Turu | Alan | Deger" formatinda.
  4) Script kapatilip tekrar acilirsa, AYNI Excel dosyasi varsa onu SILMEZ,
     mevcut satirlarin ALTINA eklemeye devam eder.

Kurulum (repo'yu YENI klonlayan biri icin - .venv/ bilerek git'e eklenmedi,
kisiye/makineye ozel oldugu icin her klonda yeniden kurulmasi gerekiyor):
  cd rp2040-original/testfiles
  python3 -m venv .venv
  ./.venv/bin/pip install -r requirements.txt

Kullanim (kurulumdan sonra, ya da zaten kuruluysa):
  ./.venv/bin/python3 field_logger.py

(Sistem Python'una hic dokunulmuyor - CachyOS/Arch gibi dagitimlarin
"externally managed environment" korumasi yuzunden, bagimliliklar bilerek
izole bir sanal ortamda tutuluyor.)
"""
import os
import sys
import time
from datetime import datetime

import serial
import serial.tools.list_ports
from openpyxl import Workbook, load_workbook

# --------------------------------------------------------------------------
# Protokol sabitleri - readout-mode.py'deki (-rm / readout mode) ile BIREBIR
# ayni. Kasitli olarak aynen kopyalandi: bu script'in tek amaci, zaten
# fiziksel olarak dogrulanmis bir protokol adimini tekrar tekrar calistirmak,
# farkli/denenmemis bir varyant riske girmeye degmez.
# --------------------------------------------------------------------------
COMM_REQUEST = b"/?!\r\n"
BAUD_RATES = [300, 600, 1200, 2400, 4800, 9600, 19200]
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "saha_test_kayitlari.xlsx")

# "Uzun/tidy" format: her tekil deger kendi satirinda - "tek tek alt alta"
# istegi tam bu yuzden boyle. Bir okuma dongusu (bir "-rm" istegi) tek basina
# 1 (duz okuma) + 10 (kisa okuma ozet alanlari) + 10 (esik kayitlari) +
# 12 (reset kayitlari) + 1 (BCC durumu) = 34 satir uretiyor.
EXCEL_COLUMNS = ["PC Tarih/Saat", "Okuma Turu", "Alan", "Deger"]

# Kisa okuma ozet alanlari - OBIS kodu -> insan-okunur isim (readout-mode.py'nin
# -rm ciktisinda hep birlikte gelen, "anlik durum" niteligindeki alanlar).
KISA_OKUMA_ALANLARI = [
    ("0.0.0", "Seri No"),
    ("0.2.0", "Firmware Versiyonu"),
    ("0.8.4", "Load Profile Periyodu"),
    ("0.9.1", "Cihaz Saati"),
    ("0.9.2", "Cihaz Tarihi"),
    ("96.1.3", "Uretim Tarihi"),
    ("96.3.12", "VRMS Esik"),
    ("32.7.0", "VRMS Max"),
    ("52.7.0", "VRMS Min"),
    ("72.7.0", "VRMS Ortalama"),
]
UZUN_OKUMA_ESIK_SLOT_SAYISI = 10   # 96.77.4*1 .. *10
UZUN_OKUMA_RESET_SLOT_SAYISI = 12  # 0.1.2*1 .. *12


def calculate_bcc(data):
    """readout-mode.py'deki calculateBCC(data, data[0]) ile ayni: ilk bayt
    hem baslangic degeri hem de dongude tekrar XOR'landigi icin kendi
    kendini iptal ediyor (X^X=0) - net etki, ilk baytin (STX) disindaki
    her seyin XOR'u."""
    xor = data[0]
    for b in data:
        xor ^= b
    return xor


def probe_port(port_name, timeout=1.5):
    """Bu portta gercek bir sayac var mi diye dener. Varsa kimlik yanitini
    (metin olarak) doner, yoksa None."""
    try:
        ser = serial.Serial(port_name, baudrate=300, bytesize=7, parity="E", stopbits=1, timeout=timeout)
    except (serial.SerialException, OSError):
        return None
    try:
        ser.reset_input_buffer()
        ser.write(COMM_REQUEST)
        time.sleep(0.3)
        resp = ser.readline()
        if len(resp) < 8 or resp[0] != 0x2F:  # 0x2F = '/'
            return None
        text = resp.decode(errors="replace").strip()
        if "<1>" not in text and "<2>" not in text:
            return None
        return text
    except Exception:
        return None
    finally:
        ser.close()


def scan_ports():
    return [(p, probe_port(p.device)) for p in serial.tools.list_ports.comports()]


def choose_port():
    while True:
        print("\nPortlar taraniyor...")
        results = scan_ports()
        if not results:
            print("Hic seri port bulunamadi. Kabloyu/adaptörü kontrol et.")
        else:
            for i, (p, info) in enumerate(results, start=1):
                if info:
                    print(f"  {i}) {p.device}  ({p.description})  -> CIHAZ BULUNDU: {info}")
                else:
                    print(f"  {i}) {p.device}  ({p.description})  -> yanit yok")

        rescan_opt = len(results) + 1
        print(f"  {rescan_opt}) Tekrar tara (kablo yeni takildiysa)")
        secim = input(f"\nHangisini kullanalim? (1-{len(results)}, ya da {rescan_opt}, q=cikis): ").strip().lower()

        if secim == "q":
            sys.exit(0)
        if secim == str(rescan_opt):
            continue
        try:
            idx = int(secim)
            if 1 <= idx <= len(results):
                return results[idx - 1][0].device
        except ValueError:
            pass
        print("Gecersiz secim, tekrar dene.")


def ask_interval_minutes():
    while True:
        raw = input("\nKac dakikada bir olcum kaydedilsin? (orn. 1, 3, 5): ").strip()
        try:
            minutes = int(raw)
            if minutes >= 1:
                return minutes
        except ValueError:
            pass
        print("Gecerli bir tam sayi gir (1 veya daha buyuk).")


def read_meter(port_name, timeout=2.0):
    """readout-mode.py'nin -rm akisinin birebir ayni protokol adimlari.
    Basarili olursa (identity_text, baud_idx, fields_dict, bcc_ok) doner:
      - identity_text: "duz okuma" kimlik/el sikisma yaniti
      - baud_idx: el sikismada anlasilan baud kodu (0-6, BAUD_RATES[baud_idx]
        gercek baud degeri) - protokolde AYRI bir OBIS alani olarak yok, sadece
        kimlik yanitinin icine gomulu, o yuzden ayrica donduruyoruz
      - fields_dict: hem "kisa okuma" (ozet alanlar) hem "uzun okuma" (esik/
        reset kayitlari) alanlarinin HEPSI - OBIS kodu -> parantez ici deger
        gruplarinin LISTESI (bazi kayitlarda "(tarih)(vrms,varyans)" gibi
        birden fazla grup oluyor)
    Tamamen basarisizsa None.

    ⚠️ Kalibrasyon sabiti burada YOK ve olamaz - bu deger RS485/IEC62056-21
    protokolunun hic bir yerinde raporlanmiyor, sadece bizim ayrica ekledigimiz
    BLE arayuzunde var. Uydurup eklemek yerine bilerek disarida birakildi."""
    # 1. asama: 300 baud'da kimlik dogrulama + mod secimi
    ser = serial.Serial(port_name, baudrate=300, bytesize=7, parity="E", stopbits=1, timeout=timeout)
    try:
        ser.reset_input_buffer()
        ser.write(COMM_REQUEST)
        time.sleep(0.25)
        resp = bytearray(ser.readline())
        time.sleep(0.25)

        if len(resp) < 8 or resp[0] != 0x2F:
            return None
        identity_text = resp.decode(errors="replace").strip()
        if "<1>" not in identity_text and "<2>" not in identity_text:
            return None

        baud_idx = int(chr(resp[4]))
        if not (0 <= baud_idx <= 6):
            return None

        info_msg = bytearray(b"\x0600\r\n")
        info_msg[2:2] = bytes([ord(str(baud_idx))])
        ser.write(info_msg)
        time.sleep(0.25)
    finally:
        ser.close()

    # 2. asama: pazarlik edilen (genelde cok daha hizli) baud'da asil veriyi oku
    ser2 = serial.Serial(port_name, baudrate=BAUD_RATES[baud_idx], bytesize=7, parity="E", stopbits=1, timeout=timeout)
    buf = bytearray()
    try:
        while True:
            line = bytearray(ser2.readline())
            if len(line) == 0:
                break
            buf += line
    finally:
        ser2.close()

    if len(buf) == 0:
        return None

    bcc_incoming = buf.pop()
    bcc_calculated = calculate_bcc(buf)
    bcc_ok = (bcc_incoming == bcc_calculated)

    text = bytes(buf).decode(errors="replace").replace("\r", "")
    fields = {}
    for line in text.split("\n"):
        line = line.strip().lstrip("\x02").rstrip("\x03").strip()
        if "(" in line and ")" in line:
            code, _, rest = line.partition("(")
            # ⚠️ Esik kayitlari gibi bazi satirlarda IKI ayri parantez grubu
            # var: "96.77.4*3(26-08-05,17:16:56)(030,23735)" - sadece ilk/son
            # parantezi almak "17:16:56)(030" gibi okunmasi zor, birlesik bir
            # deger uretiyordu. Bunun yerine TUM parantez gruplarini ayri ayri
            # tutuyoruz (fields[code] artik bir liste), asagida
            # format_field_value() bunlari duzgun birlestirip okunakli hale
            # getiriyor.
            groups = rest.rstrip(")").split(")(")
            fields[code.strip()] = [g.strip() for g in groups]

    return identity_text, baud_idx, fields, bcc_ok


def open_or_create_workbook(path):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        print(f"Mevcut kayit dosyasi bulundu, devam edilecek: {path} ({ws.max_row - 1} onceki kayit)")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Olcumler"
        ws.append(EXCEL_COLUMNS)
        print(f"Yeni kayit dosyasi olusturuldu: {path}")
    return wb, ws


DEVICE_RETRY_INTERVAL_S = 30  # cihaz (meter) yanit vermezken - adaptor hala takiliyken - bu araliklarla dene
PORT_POLL_INTERVAL_S = 5      # adaptor cikince, tekrar takilmasini bu araliklarla kontrol et


def port_exists(port_name):
    """Adaptorun kendisi (RS485 <-> USB donusturucu) hala isletim sisteminde
    goruluyor mu - katman 1 kontrolu, cihazin kendisinden BAGIMSIZ."""
    return any(p.device == port_name for p in serial.tools.list_ports.comports())


def wait_for_port_change(previous_ports):
    """Adaptor kaybolunca cagrilir - port listesi degisene kadar (yeniden
    takilana kadar) sessizce bekler. Baska, ilgisiz portlarin (orn. ESP32
    kartinin kendi USB-Serial-JTAG'i) hep takili kalmasi yuzunden yanlislikla
    hemen tetiklenmesin diye, "herhangi bir port var mi" degil, "liste
    ONCEKINDEN FARKLI mi" diye bakiyoruz."""
    while True:
        time.sleep(PORT_POLL_INTERVAL_S)
        current = {p.device for p in serial.tools.list_ports.comports()}
        if current != previous_ports:
            return


def format_datetime_pair(raw):
    """'26-08-05,18:17:06' -> '2026-08-05 18:17:06' (okunakli). Bos slot
    sentinel'i ('00-00-00,00:00:00') -> '(bos)'."""
    if not raw or not raw.strip():
        return ""
    if raw.startswith("00-00-00"):
        return "(bos)"
    date_part, _, time_part = raw.partition(",")
    return f"20{date_part} {time_part}"


def format_threshold_value(groups):
    """['26-08-05,17:16:56', '030,23735'] gibi IKI parantez grubunu
    (tarih/saat + vrms/varyans) tek, okunakli bir metinde birlestirir -
    onceki "26-08-05,18:23:47)(031,26838" gibi ham/karisik goruntunun
    duzeltilmis hali."""
    if not groups or not groups[0]:
        return ""
    dt = format_datetime_pair(groups[0])
    if dt == "(bos)" or len(groups) < 2:
        return dt
    vrms_str, _, var_str = groups[1].partition(",")
    try:
        return f"{dt}  |  VRMS={int(vrms_str)}V  Varyans={int(var_str)}"
    except ValueError:
        return f"{dt}  |  {groups[1]}"


def format_reset_value(groups):
    if not groups or not groups[0]:
        return ""
    return format_datetime_pair(groups[0])


def log_full_reading(ws, ts, identity_text, baud_idx, fields, bcc_ok):
    """Bir okuma dongusunde gelen HER seyi (duz okuma kimlik yaniti + baud
    rate + kisa okuma ozet alanlari + uzun okuma esik/reset kayitlari) tek
    tek, ayri satirlar olarak ekler. Kac satir eklendigini doner."""
    count = 0

    # --- Duz okuma: kimlik/el sikisma yaniti + ondan cikarilan baud rate ---
    # (baud rate protokolde AYRI bir OBIS alani degil, kimlik yanitinin icine
    # gomulu - okunabilir olsun diye ayri satir olarak cikariliyor)
    ws.append([ts, "Duz Okuma", "Kimlik Yaniti", identity_text])
    count += 1
    baud_deger = BAUD_RATES[baud_idx] if 0 <= baud_idx < len(BAUD_RATES) else ""
    ws.append([ts, "Duz Okuma", "Baud Rate (anlasilan)", baud_deger])
    count += 1

    # --- Kisa okuma: anlik durum ozeti alanlari ---
    for code, ad in KISA_OKUMA_ALANLARI:
        groups = fields.get(code, [])
        deger = groups[0] if groups else ""
        ws.append([ts, "Kisa Okuma", ad, deger])
        count += 1

    # --- Uzun okuma: esik asim kayitlari (bos slotlar dahil, hepsi) ---
    for i in range(1, UZUN_OKUMA_ESIK_SLOT_SAYISI + 1):
        code = f"96.77.4*{i}"
        deger = format_threshold_value(fields.get(code, []))
        ws.append([ts, "Uzun Okuma - Esik Kaydi", f"Esik Kaydi #{i}", deger])
        count += 1

    # --- Uzun okuma: reset/acilis kayitlari (bos slotlar dahil, hepsi) ---
    for i in range(1, UZUN_OKUMA_RESET_SLOT_SAYISI + 1):
        code = f"0.1.2*{i}"
        deger = format_reset_value(fields.get(code, []))
        ws.append([ts, "Uzun Okuma - Reset Kaydi", f"Reset Kaydi #{i}", deger])
        count += 1

    # --- Butunluk kontrolu (BCC) - butun dongu icin bir kere ---
    ws.append([ts, "Sistem", "BCC Kontrolu", "OK" if bcc_ok else "UYUSMADI (supheli veri)"])
    count += 1

    return count


def save_safely(wb):
    try:
        wb.save(EXCEL_PATH)
    except PermissionError:
        print("  UYARI: Excel dosyasi baska bir programda acik olabilir (orn. Microsoft Excel), "
              "bu kayit diske yazilamadi! Dosyayi kapat, bir sonraki periyotta tekrar denenecek.")


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def main_loop(port_name, interval_minutes):
    interval_s = interval_minutes * 60
    wb, ws = open_or_create_workbook(EXCEL_PATH)

    print(f"\nKayit basladi. Port: {port_name} | periyot: {interval_minutes} dk | dosya: {EXCEL_PATH}")
    print("Durdurmak icin Ctrl+C.\n")

    device_down = False  # cihaz yanit vermiyor durumu (adaptor takili ama meter cevap vermiyor)
    next_tick = time.time()  # ilk olcumu hemen al, sonrasi periyoduna gore

    while True:
        # --- KATMAN 1: RS485 adaptoru (port) hala takili mi? ---
        # Cihazin kendisiyle hic konusmadan once, once bunu kontrol ediyoruz -
        # sirali tani: once baglanti yolu (adaptor), sonra hattin ucundaki cihaz.
        if not port_exists(port_name):
            ts = now_str()
            print(f"\n[{ts}] ⚠ RS485 adaptoru '{port_name}' portunda artik bulunamiyor - cikarilmis olabilir.")
            ws.append([ts, "SISTEM", "Durum", "ADAPTOR BAGLANTISI KESILDI"])
            save_safely(wb)

            snapshot = {p.device for p in serial.tools.list_ports.comports()}
            print("Tekrar takmani bekliyorum (port listesi degisince haber verecegim)...")
            wait_for_port_change(snapshot)

            print("\nPort durumu degisti - hangi portu kullanacagimizi TEKRAR SECMEN gerekiyor "
                  "(cikarilip farkli bir isimle geri gelmis olabilir):")
            port_name = choose_port()
            device_down = False
            next_tick = time.time()
            continue

        now = time.time()
        if now < next_tick:
            time.sleep(min(next_tick - now, 5))  # 5sn'lik dilimlerle bekle, Ctrl+C'ye hizli tepki versin
            continue

        ts = now_str()

        # --- KATMAN 2: adaptor takili, ama hattin ucundaki cihaz yanit veriyor mu? ---
        try:
            result = read_meter(port_name)
        except (serial.SerialException, OSError) as e:
            print(f"[{ts}] Port erisim hatasi: {e}")
            result = None
        except Exception as e:
            print(f"[{ts}] Beklenmeyen hata: {e}")
            result = None

        if result is None:
            if not device_down:
                # Kopmanin BASLANGICINI bir kere kaydet - her 30sn'de bir tekrar
                # tekrar satir eklemiyoruz, uzun bir elektrik kesintisinde Excel
                # dosyasi binlerce "kopuk" satirla dolmasin diye.
                print(f"[{ts}] Cihazdan yanit alinamiyor - baglantisi kopmus/elektrigi kesilmis olabilir. "
                      f"Adaptor hala takili, her {DEVICE_RETRY_INTERVAL_S}sn'de bir otomatik tekrar denenecek.")
                ws.append([ts, "SISTEM", "Durum", "CIHAZ BAGLANTISI KOPTU"])
                save_safely(wb)
                device_down = True
            next_tick = time.time() + DEVICE_RETRY_INTERVAL_S
            continue

        # Basarili okuma - eger az once "kopuk" durumdaysak, donusu de ayrica belirt
        identity_text, baud_idx, fields, bcc_ok = result
        if device_down:
            print(f"[{ts}] Cihaz tekrar yanit veriyor, normal kayit periyoduna donuluyor.")
            device_down = False

        satir_sayisi = log_full_reading(ws, ts, identity_text, baud_idx, fields, bcc_ok)
        vmax = fields.get("32.7.0", [""])[0]
        vmin = fields.get("52.7.0", [""])[0]
        vmean = fields.get("72.7.0", [""])[0]
        print(f"[{ts}] Kayit alindi ({satir_sayisi} satir) - VRMS max={vmax} min={vmin} ort={vmean}")

        save_safely(wb)
        next_tick = time.time() + interval_s


def main():
    print("=== Sayac Saha Testi - Surekli Kayit ===")
    port_name = choose_port()
    interval_minutes = ask_interval_minutes()
    try:
        main_loop(port_name, interval_minutes)
    except KeyboardInterrupt:
        print("\n\nDurduruldu (Ctrl+C). Kayit dosyasi korundu, tekrar calistirinca kaldigi yerden devam eder.")


if __name__ == "__main__":
    main()
